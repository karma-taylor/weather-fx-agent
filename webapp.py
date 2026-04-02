import os
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict

import requests
from apscheduler.schedulers.background import BackgroundScheduler
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from pydantic import BaseModel, Field


FX_URL = "https://open.er-api.com/v6/latest"

# 银行偏移系数（演示用途）。真实生产建议接入各银行官方牌价接口。
BANK_RATE_MULTIPLIER: Dict[str, float] = {
    "中国银行": 1.0000,
    "中国农业银行": 0.9995,
    "中国工商银行": 1.0003,
    "中国建设银行": 1.0001,
    "威海银行": 1.0002,
}


class ConvertRequest(BaseModel):
    base: str = Field(..., min_length=3, max_length=3)
    target: str = Field(..., min_length=3, max_length=3)
    amount: float = Field(..., gt=0)
    use_fixed_rate: bool = False
    fixed_rate: float = Field(0, ge=0)
    bank_source: str = "中国银行"


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
STATIC_DIR.mkdir(parents=True, exist_ok=True)
app = FastAPI(title="Weather-FX Web App")
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
TEMPLATE_FILE = BASE_DIR / "templates" / "index.html"
VISIT_LOG_DIR = BASE_DIR / "logs"
VISIT_LOG_DIR.mkdir(parents=True, exist_ok=True)
# Windows default path requested by user. On non-Windows, fallback to local project dir.
DEFAULT_EXPORT_ROOT = (
    Path(r"E:\ai\weather-fx-agent\shuju") if os.name == "nt" else (BASE_DIR / "shuju")
)
EXPORT_ROOT = Path(os.getenv("EXPORT_ROOT_DIR", str(DEFAULT_EXPORT_ROOT)))
EXPORT_ROOT.mkdir(parents=True, exist_ok=True)
EXPORT_DIR = EXPORT_ROOT
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN", "").strip()
SCHEDULER_TIMEZONE = os.getenv("SCHEDULER_TIMEZONE", "UTC")
DAILY_EXPORT_HOUR = int(os.getenv("DAILY_EXPORT_HOUR", "23"))
DAILY_EXPORT_MINUTE = int(os.getenv("DAILY_EXPORT_MINUTE", "55"))
MONTHLY_EXPORT_HOUR = int(os.getenv("MONTHLY_EXPORT_HOUR", "0"))
MONTHLY_EXPORT_MINUTE = int(os.getenv("MONTHLY_EXPORT_MINUTE", "10"))
_scheduler = BackgroundScheduler(timezone=SCHEDULER_TIMEZONE)


def _client_ip(request: Request) -> str:
    # Prefer proxy-forwarded header when deployed behind CDN/proxy.
    xff = request.headers.get("x-forwarded-for", "").strip()
    if xff:
        return xff.split(",")[0].strip()
    xri = request.headers.get("x-real-ip", "").strip()
    if xri:
        return xri
    return request.client.host if request.client else "unknown"


def _visit_log_file() -> Path:
    return VISIT_LOG_DIR / f"visits-{datetime.utcnow().strftime('%Y-%m-%d')}.log"


def _visit_log_file_by_date(date_utc: str) -> Path:
    return VISIT_LOG_DIR / f"visits-{date_utc}.log"


def _export_file_by_date(date_utc: str) -> Path:
    return EXPORT_DIR / f"daily-visits-{date_utc}.xlsx"


def _export_file_by_month(month_utc: str) -> Path:
    return EXPORT_DIR / f"monthly-visits-{month_utc}.xlsx"


def _append_visit(ip: str, path: str) -> None:
    # Format: ISO8601_UTC\tIP\tPATH
    line = f"{datetime.utcnow().isoformat()}\t{ip}\t{path}\n"
    _visit_log_file().open("a", encoding="utf-8").write(line)


def _load_visit_rows_today():
    f = _visit_log_file()
    rows = []
    if not f.exists():
        return rows
    for line in f.read_text(encoding="utf-8").splitlines():
        parts = line.split("\t")
        if len(parts) >= 3:
            rows.append({"ts": parts[0], "ip": parts[1], "path": parts[2]})
    return rows


def _load_visit_rows_by_date(date_utc: str):
    f = _visit_log_file_by_date(date_utc)
    rows = []
    if not f.exists():
        return rows
    for line in f.read_text(encoding="utf-8").splitlines():
        parts = line.split("\t")
        if len(parts) >= 3:
            rows.append({"ts": parts[0], "ip": parts[1], "path": parts[2]})
    return rows


def _write_excel_for_date(date_utc: str) -> Path:
    rows = _load_visit_rows_by_date(date_utc)
    file_path = _export_file_by_date(date_utc)

    wb = Workbook()
    ws = wb.active
    ws.title = "visits"
    ws.append(["utc_time", "ip", "path"])
    for r in rows:
        ws.append([r["ts"], r["ip"], r["path"]])

    stats = wb.create_sheet("stats")
    counts = Counter(r["ip"] for r in rows)
    stats.append(["date_utc", date_utc])
    stats.append(["total_requests", len(rows)])
    stats.append(["unique_ips", len(counts)])
    stats.append([])
    stats.append(["ip", "count"])
    for ip, cnt in counts.most_common():
        stats.append([ip, cnt])

    wb.save(file_path)
    return file_path


def _month_days(month_utc: str):
    year, month = [int(x) for x in month_utc.split("-")]
    start = datetime(year, month, 1).date()
    next_month = datetime(year + (1 if month == 12 else 0), (1 if month == 12 else month + 1), 1).date()
    cur = start
    while cur < next_month:
        yield cur.isoformat()
        cur = cur + timedelta(days=1)


def _write_monthly_excel(month_utc: str) -> Path:
    # month_utc format: YYYY-MM
    all_rows = []
    for d in _month_days(month_utc):
        all_rows.extend(_load_visit_rows_by_date(d))

    file_path = _export_file_by_month(month_utc)
    wb = Workbook()
    ws = wb.active
    ws.title = "visits"
    ws.append(["utc_time", "ip", "path"])
    for r in all_rows:
        ws.append([r["ts"], r["ip"], r["path"]])

    stats = wb.create_sheet("stats")
    ip_counts = Counter(r["ip"] for r in all_rows)
    path_counts = Counter(r["path"] for r in all_rows)
    stats.append(["month_utc", month_utc])
    stats.append(["total_requests", len(all_rows)])
    stats.append(["unique_ips", len(ip_counts)])
    stats.append([])
    stats.append(["top_ip", "count"])
    for ip, cnt in ip_counts.most_common(100):
        stats.append([ip, cnt])
    stats.append([])
    stats.append(["top_path", "count"])
    for p, cnt in path_counts.most_common():
        stats.append([p, cnt])

    wb.save(file_path)
    return file_path


def _run_daily_export_job() -> None:
    # Export yesterday's report in UTC to avoid incomplete same-day data.
    yesterday = datetime.utcnow().date() - timedelta(days=1)
    _write_excel_for_date(yesterday.isoformat())


def _run_monthly_export_job() -> None:
    # Run on day 1, export previous month.
    today = datetime.utcnow().date()
    first_this_month = today.replace(day=1)
    last_month_last_day = first_this_month - timedelta(days=1)
    month_utc = last_month_last_day.strftime("%Y-%m")
    _write_monthly_excel(month_utc)


def _today_visit_stats():
    rows = _load_visit_rows_today()
    counts = Counter(r["ip"] for r in rows)
    return {
        "date_utc": datetime.utcnow().strftime("%Y-%m-%d"),
        "total_requests": len(rows),
        "unique_ips": len(counts),
        "ip_counts": [{"ip": ip, "count": cnt} for ip, cnt in counts.most_common()],
        "recent": rows[-30:][::-1],  # latest 30 rows, newest first
    }


def _ensure_admin_allowed(token: str = "") -> None:
    if ADMIN_TOKEN and token != ADMIN_TOKEN:
        raise HTTPException(status_code=401, detail="Unauthorized")


@app.middleware("http")
async def visit_logger(request: Request, call_next):
    response = await call_next(request)
    if request.url.path in ("/", "/api/convert"):
        _append_visit(_client_ip(request), request.url.path)
    return response


@app.on_event("startup")
def _startup_jobs():
    if not _scheduler.running:
        _scheduler.add_job(
            _run_daily_export_job,
            "cron",
            hour=DAILY_EXPORT_HOUR,
            minute=DAILY_EXPORT_MINUTE,
            id="daily_excel_export",
            replace_existing=True,
        )
        _scheduler.add_job(
            _run_monthly_export_job,
            "cron",
            day=1,
            hour=MONTHLY_EXPORT_HOUR,
            minute=MONTHLY_EXPORT_MINUTE,
            id="monthly_excel_export",
            replace_existing=True,
        )
        _scheduler.start()


@app.on_event("shutdown")
def _shutdown_jobs():
    if _scheduler.running:
        _scheduler.shutdown(wait=False)


def fetch_market_rate(base: str, target: str) -> float:
    resp = requests.get(f"{FX_URL}/{base}", timeout=12)
    resp.raise_for_status()
    data = resp.json()
    rate = data.get("rates", {}).get(target)
    if rate is None:
        raise HTTPException(status_code=400, detail=f"不支持币种对：{base}->{target}")
    return float(rate)


def fetch_historical_rates(base: str, target: str, days: int):
    end_date = datetime.utcnow().date()
    start_date = end_date - timedelta(days=days - 1)
    if base == target:
        series = []
        for i in range(days):
            d = start_date + timedelta(days=i)
            series.append({"date": d.isoformat(), "rate": 1.0})
        return series

    url = f"https://api.frankfurter.app/{start_date.isoformat()}..{end_date.isoformat()}"
    try:
        resp = requests.get(url, params={"from": base, "to": target}, timeout=12)
        resp.raise_for_status()
        data = resp.json()
        rates_map = data.get("rates") or {}
        series = []
        for date_str in sorted(rates_map.keys()):
            one_day = rates_map.get(date_str) or {}
            rate = one_day.get(target)
            if isinstance(rate, (int, float)):
                series.append({"date": date_str, "rate": float(rate)})
        if series:
            return series
    except requests.RequestException:
        # Fallback below for pairs not supported by historical provider.
        pass

    # Fallback: if historical provider has no data (or 404), use current rate
    # to build a flat historical line so the UI still updates for selected pair.
    fallback_rate = fetch_market_rate(base, target)
    return [
        {"date": (start_date + timedelta(days=i)).isoformat(), "rate": float(fallback_rate)}
        for i in range(days)
    ]


@app.get("/", response_class=HTMLResponse)
def home():
    """Serve calculator page; disable caching so template updates show without stale HTML."""
    with open(TEMPLATE_FILE, "r", encoding="utf-8") as f:
        html = f.read()
    return HTMLResponse(
        content=html,
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/healthz")
def healthz():
    return {"ok": True}


@app.get("/api/rates/history")
def rates_history(base: str, target: str, days: int = 30, bank_source: str = "中国银行"):
    base = base.upper().strip()
    target = target.upper().strip()
    if days not in (7, 30, 90):
        raise HTTPException(status_code=400, detail="days 仅支持 7 / 30 / 90。")
    if len(base) != 3 or len(target) != 3:
        raise HTTPException(status_code=400, detail="base/target 必须是 3 位币种代码。")
    if bank_source not in BANK_RATE_MULTIPLIER:
        raise HTTPException(status_code=400, detail="不支持的银行来源。")

    try:
        series = fetch_historical_rates(base, target, days)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"获取历史汇率失败：{e}") from e

    multiplier = BANK_RATE_MULTIPLIER[bank_source]
    adjusted = [{"date": p["date"], "rate": round(p["rate"] * multiplier, 6)} for p in series]
    values = [x["rate"] for x in adjusted]
    return {
        "ok": True,
        "base": base,
        "target": target,
        "days": days,
        "source": bank_source,
        "series": adjusted,
        "min_rate": min(values) if values else None,
        "max_rate": max(values) if values else None,
    }


@app.get("/api/visits/today")
def visits_today(token: str = ""):
    """
    Returns today's simple traffic stats:
    - total_requests
    - unique_ips
    - ip_counts
    """
    _ensure_admin_allowed(token)
    stats = _today_visit_stats()
    return {
        "date_utc": stats["date_utc"],
        "total_requests": stats["total_requests"],
        "unique_ips": stats["unique_ips"],
        "ip_counts": stats["ip_counts"],
    }


@app.post("/api/visits/export/today")
def export_today(token: str = ""):
    _ensure_admin_allowed(token)
    date_utc = datetime.utcnow().strftime("%Y-%m-%d")
    f = _write_excel_for_date(date_utc)
    return {"ok": True, "date_utc": date_utc, "file": str(f.name)}


@app.get("/api/visits/export/{date_utc}")
def export_by_date(date_utc: str, token: str = ""):
    _ensure_admin_allowed(token)
    f = _write_excel_for_date(date_utc)
    return FileResponse(
        path=str(f),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f.name,
    )


@app.post("/api/visits/export/monthly/{month_utc}")
def export_monthly(month_utc: str, token: str = ""):
    _ensure_admin_allowed(token)
    # Expects YYYY-MM
    if len(month_utc) != 7 or month_utc[4] != "-":
        raise HTTPException(status_code=400, detail="month_utc format should be YYYY-MM")
    f = _write_monthly_excel(month_utc)
    return {"ok": True, "month_utc": month_utc, "file": f.name, "dir": str(EXPORT_DIR)}


@app.get("/api/visits/export/monthly/{month_utc}/download")
def download_monthly(month_utc: str, token: str = ""):
    _ensure_admin_allowed(token)
    if len(month_utc) != 7 or month_utc[4] != "-":
        raise HTTPException(status_code=400, detail="month_utc format should be YYYY-MM")
    f = _write_monthly_excel(month_utc)
    return FileResponse(
        path=str(f),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f.name,
    )


@app.get("/admin/visits", response_class=HTMLResponse)
def admin_visits(token: str = ""):
    _ensure_admin_allowed(token)
    stats = _today_visit_stats()
    token_hint = "?token=YOUR_ADMIN_TOKEN" if ADMIN_TOKEN else ""

    ip_rows = "".join(
        f"<tr><td>{item['ip']}</td><td>{item['count']}</td></tr>" for item in stats["ip_counts"][:50]
    ) or "<tr><td colspan='2'>No data</td></tr>"
    recent_rows = "".join(
        f"<tr><td>{r['ts']}</td><td>{r['ip']}</td><td>{r['path']}</td></tr>" for r in stats["recent"]
    ) or "<tr><td colspan='3'>No data</td></tr>"

    return f"""
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>访问统计后台</title>
  <style>
    body {{ font-family: Arial, sans-serif; background:#f5f7fb; margin:0; padding:20px; color:#1f2937; }}
    .wrap {{ max-width: 980px; margin: 0 auto; }}
    .card {{ background:#fff; border:1px solid #d7deea; border-radius:10px; padding:16px; margin-bottom:14px; }}
    h1 {{ margin:0 0 8px; }}
    .meta {{ color:#6b7280; margin-bottom:8px; }}
    .kpi {{ display:flex; gap:12px; flex-wrap:wrap; }}
    .kpi .item {{ background:#eef2ff; border-radius:8px; padding:10px 12px; }}
    table {{ width:100%; border-collapse: collapse; font-size:14px; }}
    th, td {{ border:1px solid #e5e7eb; padding:8px; text-align:left; }}
    th {{ background:#f9fafb; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h1>访问统计后台</h1>
      <div class="meta">Date(UTC): {stats["date_utc"]} {token_hint}</div>
      <div class="kpi">
        <div class="item">总请求: <b>{stats["total_requests"]}</b></div>
        <div class="item">独立 IP: <b>{stats["unique_ips"]}</b></div>
      </div>
    </div>
    <div class="card">
      <h3>Top IP</h3>
      <table>
        <thead><tr><th>IP</th><th>Count</th></tr></thead>
        <tbody>{ip_rows}</tbody>
      </table>
    </div>
    <div class="card">
      <h3>最近访问（最多30条）</h3>
      <table>
        <thead><tr><th>UTC Time</th><th>IP</th><th>Path</th></tr></thead>
        <tbody>{recent_rows}</tbody>
      </table>
    </div>
  </div>
</body>
</html>
"""


@app.post("/api/convert")
def convert(payload: ConvertRequest):
    base = payload.base.upper().strip()
    target = payload.target.upper().strip()
    amount = float(payload.amount)

    if base == target:
        raise HTTPException(status_code=400, detail="待计算币种和目标币种不能相同。")

    if payload.use_fixed_rate:
        if payload.fixed_rate <= 0:
            raise HTTPException(status_code=400, detail="固定汇率必须大于 0。")
        rate = float(payload.fixed_rate)
        mode = "固定汇率"
        source = "手工固定汇率"
    else:
        market_rate = fetch_market_rate(base, target)
        bank = payload.bank_source.strip()
        if bank not in BANK_RATE_MULTIPLIER:
            raise HTTPException(status_code=400, detail="不支持的银行来源。")
        rate = market_rate * BANK_RATE_MULTIPLIER[bank]
        mode = "网络汇率"
        source = bank

    converted = round(amount * rate, 6)
    return {
        "ok": True,
        "mode": mode,
        "source": source,
        "base": base,
        "target": target,
        "amount": amount,
        "rate": round(rate, 6),
        "converted": converted,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "note": "银行来源为演示口径，生产请接入银行官方接口。",
    }
